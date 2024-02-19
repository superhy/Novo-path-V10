'''
@author: Yang Hu
'''

import csv
import os
import warnings

import openpyxl

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from support.env_flinc_cd45 import ENV_FLINC_CD45_U
from support.env_flinc_he import ENV_FLINC_HE_FIB, ENV_FLINC_HE_BALL_BI, \
    ENV_FLINC_HE_BALL
from support.env_flinc_he import ENV_FLINC_HE_STEA, ENV_FLINC_HE_STEA_C2
from support.env_flinc_p62 import ENV_FLINC_P62_U, ENV_FLINC_P62_BALL_BI,\
    ENV_FLINC_P62_STEA_BI, ENV_FLINC_P62_LOB_BI
from support.env_flinc_psr import ENV_FLINC_PSR_FIB
from support.env_flinc_psr import ENV_FLINC_PSR_FIB_C3


os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"




def parse_flinc_clinical_elsx(xlsx_filepath, aim_columns=['steatosis_score', 'lobular_inflammation_score',
                                                          'ballooning_score', 'fibrosis_score']):
    '''
    read the new excel file (.xlsx) by row
    parse
    '''
    wb = openpyxl.load_workbook(xlsx_filepath)
    sheet = wb['clinical_data']
    
    table_heads_dict = {}
    nb_column = sheet.max_column
    for col_id in range(nb_column):
        table_heads_dict[sheet.cell(row=1, column=col_id + 1).value] = col_id
    print('<Get table heads and their column id>: \n', table_heads_dict.items() )
    
    label_dicts = {}
    for key in aim_columns:
        label_dicts[key] = {}
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0 or row[1].value is None:
            continue
        for key in aim_columns:
            if row[table_heads_dict[key]].value != 'NA':
                label_dicts[key][row[0].value] = int(row[table_heads_dict[key]].value)
    print('<Get label dicts>, like[]: \n', label_dicts)
    
    return label_dicts

def count_clinical_labels(label_dicts, aim_label_names=['steatosis_score']):
    '''
    statistic the distribution of some clinical label
    '''

    count_dicts = []
    for i, label_name in enumerate(aim_label_names):
        annotation_dict = label_dicts[label_name]
        
        count_dict = {}
        for _, value in annotation_dict.items():
            if value not in count_dict.keys():
                count_dict[value] = 0
            else:
                count_dict[value] += 1
        print('label distribution for {}:'.format(label_name), count_dict)
        count_dicts.append(count_dict)
        
    return count_dicts

def count_flinc_stain_labels(slide_label_dict_list, stain_type, aim_label_name):
    '''
    statistic the distribution of clinical label for slides after filtering the stain (like HE)
    '''
    
    label_stat = {}
    for i, label_dict in enumerate(slide_label_dict_list):
        label = list(label_dict.items())[1][1]
        if 'Score-{}'.format(str(label)) not in label_stat.keys():
            label_stat['Score-{}'.format(str(label))] = 0
        else:
            label_stat['Score-{}'.format(str(label))] += 1
            
    labels, counts = [], []
    for item in label_stat.items():
        labels.append(item[0])
        counts.append(item[1])
    print(counts)
    
    def absolute_value(val):
        a  = np.round(val/100.*np.array(counts).sum(), 0)
        return int(a)
    
    fig = plt.figure(figsize=(5, 5))
    ax = fig.add_subplot(1, 1, 1)
    ax.pie(counts, labels=labels, autopct=absolute_value)
    ax.set_title('{}-{}'.format(stain_type, aim_label_name))
    plt.tight_layout()
    plt.savefig('{}-{}.png'.format(stain_type, aim_label_name), bbox_inches='tight')
    plt.close(fig)
    

def extract_slideid_subid_for_stain(xlsx_filepath, stain_type='HE'):
    '''
    extract the dict for {slide_id: subject_id} with specific stain type from xls metafile
    '''
    wb = openpyxl.load_workbook(xlsx_filepath)
    sheet = wb['SlideList']
    
    nb_column = sheet. max_column
    slide_column = None
    subject_id_column = None
    stain_column = None
    for col_id in range(nb_column):
        if sheet.cell(row=1, column=col_id + 1).value == 'Slide':
            slide_column = col_id
        elif sheet.cell(row=1, column=col_id + 1).value == 'SubjectID':
            subject_id_column = col_id
        elif sheet.cell(row=1, column=col_id + 1).value == 'Stain':
            stain_column = col_id
            
    if slide_column is None or subject_id_column is None or stain_column is None:
        warnings.warn('miss critical column...')
        return
    
    slideid_subid_dict = {}
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue
        if row[stain_column].value is None:
            continue
        if row[stain_column].value.find(stain_type) != -1:
            slide_id = 'Sl%03d' % row[slide_column].value
            subject_id = row[subject_id_column].value
            slideid_subid_dict[slide_id] = subject_id
    print('extracted %d slides\' subject_id' % len(slideid_subid_dict.keys()) )
    return slideid_subid_dict
            

def make_flinc_slide_label(ENV_task, label_dicts, xlsx_filepath, spec_aim_label=None, show_hv=False):
    '''
    filter the specific stain type, map the slide_id to specific label_name
    '''
    
    aim_label_name = ENV_task.TASK_NAME if spec_aim_label is None else spec_aim_label # aim_label_name is the task_name here
    stain_type = ENV_task.STAIN_TYPE 
    annotation_dict = label_dicts[aim_label_name]
    
    wb = openpyxl.load_workbook(xlsx_filepath)
    sheet = wb['SlideList']
     
    nb_column = sheet. max_column
    slide_column = None
    subject_id_column = None
    stain_column = None
    for col_id in range(nb_column):
        if sheet.cell(row=1, column=col_id + 1).value == 'Slide':
            slide_column = col_id
        elif sheet.cell(row=1, column=col_id + 1).value == 'SubjectID':
            subject_id_column = col_id
        elif sheet.cell(row=1, column=col_id + 1).value == 'Stain':
            stain_column = col_id
            
    if slide_column is None or subject_id_column is None or stain_column is None:
        warnings.warn('miss critical column...')
        return

    slide_label_dict_list = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue
        if row[stain_column].value is None:
            continue
        if row[stain_column].value.find(stain_type) != -1:
            # print(row[stain_column].value)
            slide_idstr = 'Sl%03d' % row[slide_column].value
            # print(row[subject_id_column].value, type(row[subject_id_column].value), type(row[subject_id_column].value) == str)
            if row[subject_id_column].value in annotation_dict.keys():
                slide_aim_label = annotation_dict[row[subject_id_column].value]
                slide_label_dict_list.append({'slide_id': slide_idstr, aim_label_name: slide_aim_label})
                print('loaded subject_id: {} with label: {}'.format(row[subject_id_column].value, slide_aim_label) )
            elif type(row[subject_id_column].value) == str and row[subject_id_column].value.startswith('HV'):
                slide_label_dict_list.append({'slide_id': slide_idstr, aim_label_name: -1 if show_hv else 0})
                print('loaded subject_id: {} with label: {}'.format(row[subject_id_column].value, -1 if show_hv else 0) )
                
    print('<Get slide:{}_label:{}_dict>, length:{}\n like: \n'.format(stain_type, aim_label_name, len(slide_label_dict_list)), slide_label_dict_list)
    
#     csv_test_path = 'D:/workspace/Liver-path-V10/data/FLINC/meta/HE_steatosis_score.csv'
    csv_test_path = '{}/{}_{}{}.csv'.format(ENV_task.META_FOLDER, ENV_task.STAIN_TYPE, aim_label_name, '-hv' if show_hv else '')
    csv_to_df = pd.DataFrame(slide_label_dict_list)
    csv_to_df.to_csv(csv_test_path, index=False)
    print('<Make csv annotations file at: {}>'.format(csv_test_path))
    
    return slide_label_dict_list

def combine_slide_labels_group_cx(slide_label_dict_list, groups):
    '''
    combine the labels into 3 groups, cx means c? (c2, c3, ...)
    '''
    group_label_dict, group_label_count = {}, {}
    for group_value in groups.keys():
        group_label_count[group_value] = 0
        for item in groups[group_value]: # the list of group_items
            group_label_dict[item] = group_value
    print('group mapping as: {}'.format(group_label_dict))
    
    new_slide_label_dict_list = []
    label_dist_titles = list(slide_label_dict_list[0].keys())
    label_name = label_dist_titles[1] if label_dist_titles[0] == 'slide_id' else label_dist_titles[0]
    for dict_item in slide_label_dict_list:
        org_label = int(dict_item[label_name])
        if org_label not in group_label_dict.keys():
            continue
        new_slide_label_dict_list.append({'slide_id': dict_item['slide_id'], label_name: group_label_dict[org_label]} )
        group_label_count[group_label_dict[org_label] ] += 1
    print('Combine the labels into groups, with new dict:')
    for group_label in group_label_count.keys():
        print('group label -> {}, number: {}'.format(group_label, group_label_count[group_label]) )
    return new_slide_label_dict_list

def count_flinc_stain_amount(ENV_task, xlsx_filepath):
    '''
    counting the number of slides for each staining type
    '''
    stain_type = ENV_task.STAIN_TYPE
    
    wb = openpyxl.load_workbook(xlsx_filepath)
    sheet = wb['SlideList']
     
    nb_column = sheet. max_column
    stain_column = None
    for col_id in range(nb_column):
        if sheet.cell(row=1, column=col_id + 1).value == 'Stain':
            stain_column = col_id
            
    if stain_column is None:
        warnings.warn('miss stain column...')
        return
    
    nb_stain_slides = 0
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue
        if row[stain_column].value is None:
            continue
        if row[stain_column].value.find(stain_type) != -1:
            nb_stain_slides += 1
            
    print('count stain: %s with slides: %d' % (ENV_task.STAIN_TYPE, nb_stain_slides))
    return nb_stain_slides

def extract_henning_fraction_data(slide_info_xlsx_path, henning_csv_path, stain_type):
    '''
    parse the henning's analysis results file and generate a dict for usual dataframe in this project
    '''
    # read the Excel and CSV files
    df_excel = pd.read_excel(slide_info_xlsx_path)
    df_csv = pd.read_csv(henning_csv_path)
    # fill the N/A at Level column as 'a'
    df_excel['Level'] = df_excel['Level'].fillna('a')

    # mapping SubjectID & Level to case_ID & slice
    df_excel['SubjectID'] = df_excel['SubjectID'].apply(lambda x: f'P_{x}')
    df_excel.rename(columns={'Slide': 'slide_id', 'SubjectID': 'case_ID', 'Level': 'slice'}, inplace=True)
    # filtering out the line with the corresponding stain_type
    df_excel_filtered = df_excel[df_excel['Stain'] == stain_type]

    # connect the 2 DataFrame
    merged_df = pd.merge(df_excel_filtered, df_csv, on=['case_ID', 'slice'])
    # print(merged_df)
    # print(merged_df.columns)

    # decide the results need to be extract
    stain_to_percent = {'HE': 'fat_percent', 'PSR': 'collagen_percent', 'CD45': 'CD45_percent', 'P62': 'P62_percent'}
    analysis_percent = stain_to_percent.get(stain_type, 'P62_percent')
    print('load results for: ', analysis_percent)
    # fill the N/A value as 0.0
    merged_df[analysis_percent] = merged_df[analysis_percent].fillna(0.0)

    # structure the slide_id and generate the dict to store the results
    result_dict = {}
    for _, row in merged_df.iterrows():
        slide_id = f"Sl{str(row['slide_id']).zfill(3)}"
        analysis_value = row[analysis_percent]
        result_dict[slide_id] = analysis_value

    return result_dict

def _prod_henning_percentages(ENV_task, slide_info_xlsx_name, henning_csv_name):
    '''
    generate the csv file as original percentage value
    '''
    stain_type = ENV_task.STAIN_TYPE
    slide_info_xlsx_path = os.path.join(ENV_task.META_FOLDER, slide_info_xlsx_name)
    henning_csv_path = os.path.join(ENV_task.META_FOLDER, henning_csv_name)
    
    # extract the analysis results
    analysis_dict = extract_henning_fraction_data(slide_info_xlsx_path, henning_csv_path, stain_type)
    
    stain_title_prefix_dict = {'HE': 'steatosis', 'PSR': 'fibrosis', 'CD45': 'lobular_inflammation', 'P62': 'ballooning'}
    stain_title = stain_title_prefix_dict[stain_type]
    
    # Create a DataFrame from the analysis result dictionary
    df_analysis = pd.DataFrame(list(analysis_dict.items()), columns=['slide_id', f'{stain_title}_percentage'])
    print(df_analysis)
    
    # Construct the CSV file name
    csv_file_name = f'{stain_type}_{stain_title}_pct.csv'
    
    # Write the DataFrame to a CSV file
    df_analysis.to_csv(os.path.join(ENV_task.META_FOLDER, csv_file_name), index=False)
    print(f"Henning and Marta\'s analysis results have been saved to the file: {csv_file_name}")
    
def _prod_henning_pct_grade_labels(ENV_task, slide_info_xlsx_name, henning_csv_name, label_ranges):
    '''
    generate the csv file as range grades, default using for P62_percentage from Henning
    
    Args:
        label_ranges: can be [(0, 0.2), (0.2, 1), (1, )]
            cannot have (, a) like (, 0.2), cannot be None
    '''
    stain_type = ENV_task.STAIN_TYPE
    slide_info_xlsx_path = os.path.join(ENV_task.META_FOLDER, slide_info_xlsx_name)
    henning_csv_path = os.path.join(ENV_task.META_FOLDER, henning_csv_name)
    # extract the analysis results
    analysis_dict = extract_henning_fraction_data(slide_info_xlsx_path, henning_csv_path, stain_type)
    
    stain_title_prefix_dict = {'HE': 'steatosis', 'PSR': 'fibrosis', 'CD45': 'lobular_inflammation', 'P62': 'ballooning'}
    stain_title = stain_title_prefix_dict[stain_type]
    
    df_labels = pd.DataFrame(columns=['slide_id', f'{stain_title}_percentage_label'])
    for slide_id, percentage in analysis_dict.items():
        label = None
        for i, range_ in enumerate(label_ranges):
            if len(range_) == 2 and range_[0] <= percentage <= range_[1]:
                label = i
                break
            elif len(range_) == 1 and percentage > range_[0]:
                label = i
                break
        
        # setup the default label
        if label is None:
            label = len(label_ranges) - 1
        
        df_labels = df_labels.append({'slide_id': slide_id, f'{stain_title}_percentage_label': label}, ignore_index=True)
    
    print(df_labels)
    # Construct the CSV file name
    csv_file_name = f'{stain_type}_{stain_title}_pct_lbl.csv'
    
    # Write the DataFrame to a CSV file
    df_labels.to_csv(os.path.join(ENV_task.META_FOLDER, csv_file_name), index=False)
    print(f"Henning and Marta\'s analysis grade labels have been saved to the file: {csv_file_name}")
    
def load_percentages_from_csv(ENV_task, percentage_csv_name=None):
    '''
    load back Henning and Marta's percentage results from csv file
    '''
    stain_title_prefix_dict = {'HE': 'steatosis', 'PSR': 'fibrosis', 'CD45': 'lobular_inflammation', 'P62': 'ballooning'}
    stain_type = ENV_task.STAIN_TYPE
    stain_title = stain_title_prefix_dict[stain_type]
    
    if percentage_csv_name is None:
        percentage_csv_name = f'{stain_type}_{stain_title}_percentage.csv'
        
    df = pd.read_csv(os.path.join(ENV_task.META_FOLDER, percentage_csv_name))
    results_dict = pd.Series(df[f'{stain_title}_percentage'].values, index=df['slide_id']).to_dict()
    return results_dict

def get_slide_ids_with_b_cpt_lower_than_thd(csv_file_path, threshold=0.05):
    '''
    get slide_ids for those with pct_value lower than threshold
    '''
    df = pd.read_csv(csv_file_path)
    filtered_df = df[df['ballooning_percentage'] < threshold]
    slide_ids = filtered_df['slide_id'].tolist()
    
    return slide_ids
    
def _load_clinical_labels(show_hv):
    
    # TASK_ENVS = [ENV_FLINC_HE_STEA, ENV_FLINC_HE_FIB, ENV_FLINC_PSR_FIB]
    # TASK_ENVS = [ENV_FLINC_HE_BALL]
    # TASK_ENVS = [ENV_FLINC_P62_BALL_BI]
    TASK_ENVS = [ENV_FLINC_P62_BALL_BI, ENV_FLINC_P62_STEA_BI, ENV_FLINC_P62_LOB_BI]
    
    xlsx_path_clinical = '{}/FLINC_clinical_data_DBI_2022-0715_EDG.xlsx'.format(TASK_ENVS[0].META_FOLDER)
    clinical_label_dicts = parse_flinc_clinical_elsx(xlsx_path_clinical)
    _ = count_clinical_labels(clinical_label_dicts, aim_label_names=['steatosis_score'])
    
    xlsx_path_slide_1 = '{}/FLINC_23910-157_withSubjectID.xlsx'.format(TASK_ENVS[0].META_FOLDER)
    xlsx_path_slide_2 = '{}/FLINC_23910-158_withSubjectID.xlsx'.format(TASK_ENVS[0].META_FOLDER)
    # xlsx_path_slide_list = [xlsx_path_slide_1, xlsx_path_slide_1, xlsx_path_slide_1, xlsx_path_slide_2, xlsx_path_slide_2]
    xlsx_path_slide_list = [xlsx_path_slide_2, xlsx_path_slide_2, xlsx_path_slide_2]
    
    for i, task_env in enumerate(TASK_ENVS):
        slide_label_dict_list = make_flinc_slide_label(task_env, clinical_label_dicts,
                                                       xlsx_filepath=xlsx_path_slide_list[i],
                                                       show_hv=show_hv)
        count_flinc_stain_labels(slide_label_dict_list, task_env.STAIN_TYPE, task_env.TASK_NAME)
        
def _load_lobular_clinical_labels():
    
    ENV_task = ENV_FLINC_CD45_U
    xlsx_path_clinical = '{}/FLINC_clinical_data_DBI_2022-0715_EDG.xlsx'.format(ENV_task.META_FOLDER)
    clinical_label_dicts = parse_flinc_clinical_elsx(xlsx_path_clinical)
    
    xlsx_path_slide = '{}/FLINC_23910-158_withSubjectID.xlsx'.format(ENV_task.META_FOLDER)
    aim_label = 'lobular_inflammation_score'
    
    slide_label_dict_list = make_flinc_slide_label(ENV_task, clinical_label_dicts,
                                                   xlsx_filepath=xlsx_path_slide,
                                                   spec_aim_label=aim_label)
    count_flinc_stain_labels(slide_label_dict_list, ENV_task.STAIN_TYPE, aim_label)
    
def _load_ballooning_clinical_labels():
    
    ENV_task = ENV_FLINC_P62_U
    xlsx_path_clinical = '{}/FLINC_clinical_data_DBI_2022-0715_EDG.xlsx'.format(ENV_task.META_FOLDER)
    clinical_label_dicts = parse_flinc_clinical_elsx(xlsx_path_clinical)
    
    xlsx_path_slide = '{}/FLINC_23910-158_withSubjectID.xlsx'.format(ENV_task.META_FOLDER)
    aim_label = 'ballooning_score'
    
    slide_label_dict_list = make_flinc_slide_label(ENV_task, clinical_label_dicts,
                                                   xlsx_filepath=xlsx_path_slide,
                                                   spec_aim_label=aim_label)
    count_flinc_stain_labels(slide_label_dict_list, ENV_task.STAIN_TYPE, aim_label)
        
def _prod_combine_labels():
    '''
    produce the combined label csv file from ENV_task, for C_ENV_task
    '''
    ENV_task = ENV_FLINC_PSR_FIB
    C_ENV_task = ENV_FLINC_PSR_FIB_C3
    groups = {0: [0, 1], 1: [2], 2: [3, 4]}
    # ENV_task = ENV_FLINC_HE_STEA
    # C_ENV_task = ENV_FLINC_HE_STEA_C2
    # groups = {0: [0, 1], 1: [2, 3]}
    
    slide_label_dict_list = query_task_label_dict_list_fromcsv(ENV_task)
    new_slide_label_dict_list = combine_slide_labels_group_cx(slide_label_dict_list, groups=groups)
    
    csv_test_path = '{}/{}_{}.csv'.format(C_ENV_task.META_FOLDER, C_ENV_task.STAIN_TYPE, C_ENV_task.TASK_NAME)
    csv_to_df = pd.DataFrame(new_slide_label_dict_list)
    csv_to_df.to_csv(csv_test_path, index=False)
    print('<Make combined csv annotations file at: {}>'.format(csv_test_path))
    return new_slide_label_dict_list

def _prod_bi_label_combine_labels(ENV_task, groups, aim_label):
    '''
    produce binary label for flexible annotations
    
    Args:
        there are some examples for ENV_task, groups, aim_label in lobular_inflammation
            ENV_task = ENV_FLINC_CD45_U
            # groups = {0: [0], 1: [1, 2, 3]} # just for binary
            groups = {0: [0], 1:[3]} # for dual-poles
            aim_label = 'lobular_inflammation_score'
    '''
    
    slide_label_dict_list = query_task_label_dict_list_fromcsv(ENV_task, 
                                                               task_csv_filename='{}_{}.csv'.format(ENV_task.STAIN_TYPE, aim_label))
    new_slide_label_dict_list = combine_slide_labels_group_cx(slide_label_dict_list, groups)
    # 'bi' same with 'c2'
    show_hv = True if groups[0][0] == -1 else False
    aim_name = aim_label.replace('-hv', '_hv') if show_hv else aim_label + '_bi'
    csv_test_path = '{}/{}_{}.csv'.format(ENV_task.META_FOLDER, ENV_task.STAIN_TYPE, aim_name)
    csv_to_df = pd.DataFrame(new_slide_label_dict_list)
    csv_to_df.to_csv(csv_test_path, index=False)
    print('<Make combined csv annotations file at: {}>'.format(csv_test_path))
    return new_slide_label_dict_list

def _prod_bi_lobular_combine_labels():
    '''
    produce binary label file specifically for lobular inflammation
    '''
    
    ENV_task = ENV_FLINC_CD45_U
    # groups = {0: [0], 1: [1, 2, 3]} # just for binary
    groups = {0: [0], 1:[3]} # for dual-poles
    aim_label = 'lobular_inflammation_score'
    
    slide_label_dict_list = query_task_label_dict_list_fromcsv(ENV_task, 
                                                               task_csv_filename='{}_{}.csv'.format(ENV_task.STAIN_TYPE, aim_label))
    new_slide_label_dict_list = combine_slide_labels_group_cx(slide_label_dict_list, groups)
    # 'bi' same with 'c2'
    csv_test_path = '{}/{}_{}.csv'.format(ENV_task.META_FOLDER, ENV_task.STAIN_TYPE, aim_label+'_bi')
    csv_to_df = pd.DataFrame(new_slide_label_dict_list)
    csv_to_df.to_csv(csv_test_path, index=False)
    print('<Make combined csv annotations file at: {}>'.format(csv_test_path))
    return new_slide_label_dict_list

def _prod_bi_ballooning_combine_labels():
    '''
    produce binary label file specifically for lobular inflammation
    '''
    
    ENV_task = ENV_FLINC_P62_U
    # groups = {0: [0], 1: [1, 2, 3]} # just for binary
    groups = {0: [0], 1:[2]} # for dual-poles
    aim_label = 'ballooning_score'
    
    slide_label_dict_list = query_task_label_dict_list_fromcsv(ENV_task, 
                                                               task_csv_filename='{}_{}.csv'.format(ENV_task.STAIN_TYPE, aim_label))
    new_slide_label_dict_list = combine_slide_labels_group_cx(slide_label_dict_list, groups)
    # 'bi' same with 'c2'
    csv_test_path = '{}/{}_{}.csv'.format(ENV_task.META_FOLDER, ENV_task.STAIN_TYPE, aim_label+'_bi')
    csv_to_df = pd.DataFrame(new_slide_label_dict_list)
    csv_to_df.to_csv(csv_test_path, index=False)
    print('<Make combined csv annotations file at: {}>'.format(csv_test_path))
    return new_slide_label_dict_list
        
def _count_stain_amount():
    
    TASK_ENVS = [ENV_FLINC_HE_STEA, ENV_FLINC_HE_FIB, ENV_FLINC_PSR_FIB]
    
    xlsx_path_slide_1 = '{}/FLINC_23910-157_withSubjectID.xlsx'.format(TASK_ENVS[0].META_FOLDER)
    xlsx_path_slide_2 = '{}/FLINC_23910-158_withSubjectID.xlsx'.format(TASK_ENVS[0].META_FOLDER)
    xlsx_path_slide_list = [xlsx_path_slide_1, xlsx_path_slide_1, xlsx_path_slide_1, xlsx_path_slide_2, xlsx_path_slide_2]
    
    for i, task_env in enumerate(TASK_ENVS):
        _ = count_flinc_stain_amount(task_env, xlsx_filepath=xlsx_path_slide_list[i])
        

def auto_find_task_csvname(ENV_task):
    '''
    auto detect the task_csv_filename
    '''
    f_start_string = '{}_{}'.format(ENV_task.STAIN_TYPE, ENV_task.TASK_NAME)
    for f in os.listdir(ENV_task.META_FOLDER):
        if f.startswith(f_start_string) and f.endswith('.csv'):
            print('automatically find CSV file: {}'.format(f))
            task_csv_filename = f
            break
    return task_csv_filename

def query_task_label_dict_list_fromcsv(ENV_task, task_csv_filename=None):
    '''
    load the task label dict list (to maintain the sort of keys) from csv
    '''
    if task_csv_filename == None:
        task_csv_filename = auto_find_task_csvname(ENV_task)
    
    task_csv_filepath = os.path.join(ENV_task.META_FOLDER, task_csv_filename)
    slide_label_dict_list = []
    column_1, column_2 = 'slide_id', ''
    with open(task_csv_filepath, 'r', newline='') as task_csv_file:
        csv_reader = csv.reader(task_csv_file)
        for l, csv_line in enumerate(csv_reader):
            if l == 0: # skip the first line for title
                column_1, column_2 = csv_line[0], csv_line[1]
            else:
                slide_label_dict_list.append({column_1: csv_line[0], column_2: csv_line[1]})
    return slide_label_dict_list
        
def query_task_label_dict_fromcsv(ENV_task, task_csv_filename=None):
    '''
    load the task label dict from csv
    '''
    if task_csv_filename == None:
        task_csv_filename = auto_find_task_csvname(ENV_task)
            
    task_csv_filepath = os.path.join(ENV_task.META_FOLDER, task_csv_filename)
    task_label_dict = {}
    with open(task_csv_filepath, 'r', newline='') as task_csv_file:
        csv_reader = csv.reader(task_csv_file)
        for l, csv_line in enumerate(csv_reader):
            if l == 0: # skip the first line for title
                continue
            task_label_dict[csv_line[0]] = int(csv_line[1])
            
    return task_label_dict

def trans_slide_label_dict_to_subid(slideid_label_dict, slideid_subid_dict):
    '''
    transfer the slide_label_dict to using subject_id (clinical_id) as the key
    {slide_id (Sl???): label} -> {subject_id (?): label}
    '''
    subid_label_dict = {}
    for slide_id in slideid_label_dict.keys():
        sub_id = slideid_subid_dict[slide_id]
        subid_label_dict[sub_id] = slideid_label_dict[slide_id]
        
    return subid_label_dict


if __name__ == '__main__':
    ''' label preparing steps '''
    # _load_clinical_labels(show_hv=False) # this will be the 1st step
    # _load_clinical_labels(show_hv=True)
    # _count_stain_amount()
    # _prod_combine_labels()
    
    # _load_lobular_clinical_labels()
    # _ = _prod_bi_lobular_combine_labels()

    # _load_ballooning_clinical_labels()
    # _ = _prod_bi_ballooning_combine_labels()
    
    ''' here will be the 2nd step '''
    # pkg_param_fib = (ENV_FLINC_PSR_FIB, {0: [0], 1:[4]}, 'fibrosis_score')
    # pkg_param_stea = (ENV_FLINC_HE_STEA, {0: [0], 1:[3]}, 'steatosis_score')
    # pkg_param_ball = (ENV_FLINC_HE_BALL_BI, {0: [0], 1:[2]}, 'ballooning_score')
    # pkg_param_stea = (ENV_FLINC_P62_STEA_BI, {0: [0], 1:[3]}, 'steatosis_score')
    # pkg_param_lob = (ENV_FLINC_P62_LOB_BI, {0: [0], 1:[3]}, 'lobular_inflammation_score')
    # _ = _prod_bi_label_combine_labels(pkg_param_fib[0], pkg_param_fib[1], pkg_param_fib[2])
    # _ = _prod_bi_label_combine_labels(pkg_param_stea[0], pkg_param_stea[1], pkg_param_stea[2])
    # _ = _prod_bi_label_combine_labels(pkg_param_ball[0], pkg_param_ball[1], pkg_param_ball[2])
    # _ = _prod_bi_label_combine_labels(pkg_param_lob[0], pkg_param_lob[1], pkg_param_lob[2])
    
    # pkg_param_ball_hv = (ENV_FLINC_P62_BALL_BI, {0: [-1], 1:[2]}, 'ballooning_score-hv')
    # pkg_param_stea_hv = (ENV_FLINC_P62_STEA_BI, {0: [-1], 1:[3]}, 'steatosis_score-hv')
    pkg_param_lob_hv = (ENV_FLINC_P62_LOB_BI, {0: [-1], 1:[3]}, 'lobular_inflammation_score-hv')
    # _ = _prod_bi_label_combine_labels(pkg_param_ball_hv[0], pkg_param_ball_hv[1], pkg_param_ball_hv[2])
    # _ = _prod_bi_label_combine_labels(pkg_param_stea_hv[0], pkg_param_stea_hv[1], pkg_param_stea_hv[2])
    # _ = _prod_bi_label_combine_labels(pkg_param_lob_hv[0], pkg_param_lob_hv[1], pkg_param_lob_hv[2])
    
#     print(ENV_FLINC_HE_STEA.PROJECT_NAME)

    
    _prod_henning_percentages(ENV_task=ENV_FLINC_P62_U, 
                                    slide_info_xlsx_name='FLINC_23910-158_withSubjectID.xlsx', 
                                    henning_csv_name='flinc_quantitative_pathology - add Marta results.csv')
    
    _prod_henning_pct_grade_labels(ENV_task=ENV_FLINC_P62_U, 
                                    slide_info_xlsx_name='FLINC_23910-158_withSubjectID.xlsx', 
                                    henning_csv_name='flinc_quantitative_pathology - add Marta results.csv',
                                    label_ranges=[(0, 0.2), (0.2, 1), (1, )])

